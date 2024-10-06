import subprocess
import sys

def install(package):
    subprocess.check_call([sys.executable, "-m", "pip", "install", package])

install('pandas')
install('numpy')
install('chardet')
install('xlsxwriter')
install('openpyxl')


import pandas as pd
import numpy as np
import chardet 
import xlsxwriter
import re
import openpyxl
from openpyxl.utils import get_column_letter


file = r"file.csv"
def detect_encoding(file_path): 
    with open(file_path, 'rb') as file: 
        detector = chardet.universaldetector.UniversalDetector() 
        for line in file: 
            detector.feed(line) 
            if detector.done: 
                break
        detector.close() 
    return detector.result['encoding'] 

encoding = detect_encoding(file) 
print(f'The encoding of the file is: {encoding}') 


df = pd.read_csv(file, na_filter=False, sep=';', encoding= 'ISO-8859-1')
df = df.drop('Nombre', axis=1) 
df = df.drop('¿ Qué carrera profesional, técnica o tecnológica le gustaría estudiar una vez finalizada su etapa escolar?', axis=1)
df = df.drop('Correo electrónico', axis=1) 
df = df.drop('Correo electrónico2', axis=1) 
df = df.drop('ID', axis=1) 
df = df.drop('Hora de inicio', axis=1) 
df = df.drop('Hora de finalización', axis=1) 
df = df.drop('Curso', axis=1) 
df = df.drop('Colegio al cual pertenecen', axis=1) 


df.head(10)


seminars = len(df['Ordene los seminarios desde el que mas llamo su atención al que menos atrajo su interés.'].iloc[0].split(';'))
list_seminars = [f"Seminario {i}" for i in range(1, seminars +1)]
df[list_seminars] = df['Ordene los seminarios desde el que mas llamo su atención al que menos atrajo su interés.'].str.split(';', n=seminars-1, expand=True)
df = df.drop(columns=[list_seminars[-1]])
df = df.drop('Ordene los seminarios desde el que mas llamo su atención al que menos atrajo su interés.', axis=1) 
df.head()


student_preferences = {}
for index, row in df.iterrows():
    nombre_completo = row['Nombre Completo']
    seminar_preferences = [v for v in row[1:] if v]
    student_preferences[nombre_completo] = seminar_preferences
    
print (student_preferences)


total_seminars = seminars - 1
print(f'Total Seminarios: {total_seminars}')

total_students = len(df.index)
print(f'\nNumero de estudiantes: {total_students}')

students_seminar =  np.ceil(total_students/total_seminars)
print(f'\nNumero de estudiantes por seminario: {students_seminar}')


seminar_capacities = {}
for student, preferences in student_preferences.items():
    for seminar in preferences:
        if seminar not in seminar_capacities:
            seminar_capacities[seminar] = students_seminar
            
print (seminar_capacities)


def assign_seminars(student_preferences, seminar_capacities):
    assigned_seminars = {}
    for student, preferences in student_preferences.items():
        for seminar in preferences:
            if seminar_capacities[seminar] > 0:
                assigned_seminars[student] = seminar
                seminar_capacities[seminar] -= 1
                break
    return assigned_seminars

assigned_seminars = assign_seminars(student_preferences, seminar_capacities)
print(assigned_seminars)


final_df = pd.DataFrame(list(assigned_seminars.items()), columns=['Nombre', 'Seminario'])

final_df.head()


workbook = xlsxwriter.Workbook('asignacion_seminarios.xlsx')
seminarios = final_df['Seminario'].unique()

for seminario in seminarios:
    sheet_name = re.sub(r'[:?/\\*\[\]\'"]', '_', seminario)[:31]
    worksheet = workbook.add_worksheet(sheet_name)
    worksheet.set_default_row(15)
    seminario_df = final_df[final_df['Seminario'] == seminario].dropna(how='all')
    worksheet.write_row(0, 0, seminario_df.columns.tolist())
    for index, row in seminario_df.iterrows():
        worksheet.write_row(index + 1, 0, row.tolist())

workbook.close()


wb = openpyxl.load_workbook('asignacion_seminarios.xlsx')
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    for row in range(ws.max_row, 0, -1):
        if all(cell.value is None for cell in ws[row]):
            ws.delete_rows(row)


wb.save('asignacion_seminarios.xlsx')


